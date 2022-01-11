from openpyxl import Workbook, load_workbook
import time

def main():

    #Login de Usuário
    usuario = str(input("Qual o seu usuário?\n")).lower()

    '''#Coletar programação das obras
    Programação = load_workbook("C:\\Users\\{}\\Documents\\Documentos Obras\\Programação\\PROGRAMADOR_CONTROL_01.xlsx".format(usuario))
    Aba_programação = Programação["PROGRAMAÇÃO "]
    #Planilha de saida
    saida = load_workbook("C:\\Users\\{}\\Documents\\Viagem Neopolis.xlsx".format(usuario))
    Aba_PC = saida.active'''
    
    #Coletar informações dos colaboradores na base de dados
    Planilha_base_de_dados = load_workbook("C:\\Users\\{}\\Documents\\Colaboradores\\Colaboradores.xlsx".format(usuario))
    Aba = Planilha_base_de_dados.active
    #Planilha de Justificativa
    
    #Txt com a justificativa
    caminho_saida = open("C:\\Users\\{}\\Documents\\Viagem Neopolis.txt".format(usuario), 'w') # Saída do resultado

    #Base de dados Colaboradores
    funcionários = []
    matriculas = []
    funções = []
    equipes= []

    
    for celula_nome in Aba['B']:  
        linha_nome = celula_nome.row
        nome = str(Aba["B{}".format(linha_nome)].value)
        if nome == "Funcionário":
            time.sleep(0.00001)
        else:
            funcionários.append(nome)

    for celula_matricula in Aba['A']:  
        linha_matricula = celula_matricula.row
        matricula = str(Aba["A{}".format(linha_matricula)].value)
        if matricula == "Matricula":
            time.sleep(0.00001)
        else:
            matriculas.append(matricula)

    for celula_função in Aba['C']:  
        linha_função= celula_função.row
        função = str(Aba["C{}".format(linha_função)].value)
        if função == "Função":
            time.sleep(0.00001)
        else:
            funções.append(função)

    for celula_equipe in Aba['D']:  

        linha_equipe = celula_equipe.row
        equipe = str(Aba["D{}".format(linha_equipe)].value)
        if equipe == "Equipe":
            time.sleep(0.00001)
        else:
          equipes.append(equipe)

    #Escrever Informação
    '''Data_inicial = int(input("Qual a data inicial\n"))
    Data_Final = int(input("Qual a data final?\n"))
    mes = int(input("Qual o mês?\n"))
    ano = int(input("Qual o ano?\n"))
    linha_inicial = int(input("Qual o nº da linha inicial da programação?\n"))

    dia = Data_inicial

    datacompleta = str('{}/{}/{}'.format(dia, mes, ano))
    
    for celula in Aba['C']:
        linha = celula.row
        
        if linha < linha_inicial:
            time.sleep(0.0000001)
        else:
            data = str(Aba_programação["C{}".format(linha)].value)
            obra = str(Aba_programação["D{}".format(linha)].value)
            PES = str(Aba_programação["W{}".format(linha)].value)
            EqpP = str(Aba_programação["H{}".format(linha)].value)
            EqpS = str(Aba_programação["I{}".format(linha)].value)
            HI = str(Aba_programação["K{}".format(linha)].value)
            HT = str(Aba_programação["L{}".format(linha)].value)
            local = str(Aba_programação["AC{}".format(linha)].value)
            Endereço = str(Aba_programação["AD{}".format(linha)].value)

            if data == datacompleta:
                linhapc = (linha-linha_inicial)+2
                Aba_PC["A{}".format(linhapc)] = data
                Aba_PC["B{}".format(linhapc)] = obra
                Aba_PC["C{}".format(linhapc)] = PES
                Aba_PC["D{}".format(linhapc)] = EqpP
                Aba_PC["E{}".format(linhapc)] = EqpS
                Aba_PC["F{}".format(linhapc)] = HI
                Aba_PC["G{}".format(linhapc)] = HT
                Aba_PC["F{}".format(linhapc)] = local
                Aba_PC["G{}".format(linhapc)] = Endereço

                dia = dia+1
            
            elif data == "21/01/2022" and local == "SEM LOTE":
                linhapc = (linha-linha_inicial)+2
                Aba_PC["A{}".format(linhapc)] = data
                Aba_PC["B{}".format(linhapc)] = obra
                Aba_PC["C{}".format(linhapc)] = PES
                Aba_PC["D{}".format(linhapc)] = EqpP
                Aba_PC["E{}".format(linhapc)] = EqpS
                Aba_PC["F{}".format(linhapc)] = HI
                Aba_PC["G{}".format(linhapc)] = HT
                Aba_PC["H{}".format(linhapc)] = "ITAPORANGA D'AJUDA"
                Aba_PC["I{}".format(linhapc)] = 'POVOADO DURO'
                dia = dia+1'''

    EQUIPES_VIAGEM = ['LM 01', 'LM 02', 'LM 03', 'LV 01']
    
    for grupo in EQUIPES_VIAGEM:
        
        if grupo == "LM 02":
            caminho_saida.write('Solicito reserva para os seguintes colaboradores para cumprir com programação de Obras na região de Neópolis e arredores, no período de 17/01/2022 à 20/01/2022\n')
            indice = 0
            while indice < len(equipes):
                if equipes[indice] == grupo:
                    caminho_saida.write('Nome: {}\nMatrícula: {}\nCargo: {}\n\n'.format(funcionários[indice], matriculas[indice], funções[indice]))
                    indice = indice+1
                else:
                    indice = indice+1
        
        else:
            caminho_saida.write('Boa Tarde,\nSolicito reserva para os seguintes colaboradores para cumprir com programação de Obras na região de Neópolis e arredores, no período de 17/01/2022 à 21/01/2022\n')             
            indice = 0
            while indice < len(equipes):
                if equipes[indice] == grupo:
                    caminho_saida.write('Nome: {}\nMatrícula: {}\nCargo: {}\n\n'.format(funcionários[indice], matriculas[indice], funções[indice]))
                    indice = indice+1
                else:
                    indice = indice+1

    #saida.save("C:\\Users\\{}\\Documents\\Viagem Neopolis.xlsx".format(usuario))
    
    caminho_saida.close()

main()